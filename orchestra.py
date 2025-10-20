"""
===============================================================================
 Project Name:    Complete Master Orchestrator - Final Synchronized Workflow
 File Name:       Master_Orchestrator_Final_Compiled_Fixed.py
 Description:     Runs 7 interconnected processes using external index and caching.
 Author:          Akshay Solanki (Compiled Final Version with Fixes)
 Created on:      21-Oct-2025
 Dependencies:    pandas, openpyxl, tkinter, tqdm, fitz, PyPDF2, re, hashlib, json
===============================================================================
"""

import os
import shutil
import pandas as pd
import re
import hashlib
from datetime import datetime
from collections import defaultdict
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tqdm import tqdm
import time
import fitz        # Used for P6 PDF merging and content hashing
import PyPDF2    # Used for P3 simple page extraction
import json

# --- GLOBAL CONSTANTS ---
LOG_FILE = os.path.join(os.getcwd(), "orchestrator_log.txt")
ERROR_REPORT = os.path.join(os.getcwd(), "error_report.txt")
P4_CACHE_FILE = os.path.join(os.getcwd(), "linewise_index_cache.json")
P6_CACHE_NAME = ".merge_cache.json"
PDF_INDEX_REFERENCE_NAME = "PDF_TOC_Reference.txt"
OUTPUT_EXCEL_NAME = "output.xlsx" # System-level Excel file name
FINAL_PDF_SUFFIX = ".pdf"

# --- UTILITY FUNCTIONS ---
def log_msg(msg):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{timestamp}] {msg}"
    print(line)
    try:
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(line + "\n")
    except Exception as e:
        print(f"ERROR logging: {e}")

def log_error(msg):
    with open(ERROR_REPORT, "a", encoding="utf-8") as f:
        f.write(msg + "\n")

def select_folder(title):
    root = tk.Tk()
    root.withdraw()
    folder = filedialog.askdirectory(title=title)
    root.destroy()
    return folder

def select_file(title, filetypes):
    root = tk.Tk()
    root.withdraw()
    file = filedialog.askopenfilename(title=title, filetypes=filetypes)
    root.destroy()
    return file

def format_time(seconds):
    hrs = int(seconds // 3600)
    mins = int((seconds % 3600) // 60)
    secs = int(seconds % 60)
    return f"{hrs:02d}:{mins:02d}:{secs:02d}"

# --- PROCESS 1: ISO MANAGER ---
def safe_copy(src, dest):
    if not os.path.exists(src):
        raise FileNotFoundError(src)
    dest_dir = os.path.dirname(dest)
    os.makedirs(dest_dir, exist_ok=True)
    base, ext = os.path.splitext(dest)
    if os.path.exists(dest) and os.path.getsize(src) == os.path.getsize(dest):
        return dest
    candidate = dest
    i = 1
    while os.path.exists(candidate):
        candidate = f"{base}_dup{i}{ext}"
        i += 1
    shutil.copy2(src, candidate)
    return candidate

def make_folder_name(loop_no: str, system_no: str) -> str:
    return f"{str(loop_no).strip()}_{str(system_no).strip()}"

def find_iso_on_server(iso_no: str, server_path: str) -> str:
    if not iso_no or not isinstance(iso_no, str):
        return ""
    iso_no = iso_no.strip()
    if os.path.isdir(server_path):
        try:
            for item in os.listdir(server_path):
                if item.lower().endswith(FINAL_PDF_SUFFIX):
                    if "(" in item and ")" in item:
                        start = item.rfind("(")
                        end = item.rfind(")")
                        if start < end:
                            extracted = item[start+1:end]
                            if extracted.lower() == iso_no.lower():
                                full_path = os.path.join(server_path, item)
                                if os.path.isfile(full_path):
                                    return full_path
        except PermissionError:
            pass
    return ""

def create_or_update_excel(excel_file):
    headers = ["Iso no", "loop no", "system no", "folder name", "history folder name", "ISO Status"]
    if not os.path.exists(excel_file):
        df = pd.DataFrame(columns=headers)
        df.to_excel(excel_file, index=False)
        log_msg(f"Created Excel: {excel_file}")
        messagebox.showinfo("Excel Created", f"Fill first three columns and save.")
        return False
    df = pd.read_excel(excel_file, dtype=str).fillna("")
    for col in headers:
        if col not in df.columns:
            df[col] = ""
    df = df[headers]
    for idx, row in df.iterrows():
        new_folder_name = make_folder_name(row["loop no"], row["system no"])
        df.at[idx, "folder name"] = new_folder_name
        if not row["history folder name"]:
            df.at[idx, "history folder name"] = new_folder_name
    df.to_excel(excel_file, index=False)
    return True

def highlight_missing_iso(excel_file):
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        clear_fill = PatternFill(fill_type=None)
        header = [cell.value for cell in ws[1]]
        if "ISO Status" not in header:
            return
        col_idx = header.index("ISO Status") + 1
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            status_val = str(cell.value).strip().upper() if cell.value else ""
            fill = red_fill if status_val == "MISSING" else clear_fill
            for c in range(1, ws.max_column + 1):
                ws.cell(row=row, column=c).fill = fill
        wb.save(excel_file)
    except Exception as e:
        log_msg(f"ERROR highlighting: {e}")

def iso_manager(excel_file, server_path, dest_root):
    process_start = time.time()
    log_msg("=== P1: ISO MANAGER START ===")
    
    df = pd.read_excel(excel_file, dtype=str).fillna("")
    for idx, row in df.iterrows():
        df.at[idx, "folder name"] = make_folder_name(row["loop no"], row["system no"])

    processed_folders = {}
    for idx, row in df.iterrows():
        desired = row["folder name"].strip()
        history = row["history folder name"].strip()
        if not desired or not history:
            continue
        key = (history, desired)
        if key in processed_folders:
            df.at[idx, "history folder name"] = desired
            continue
        desired_path = os.path.join(dest_root, desired)
        history_path = os.path.join(dest_root, history)
        try:
            if history != desired and os.path.exists(history_path):
                if not os.path.exists(desired_path):
                    os.rename(history_path, desired_path)
                    log_msg(f"RENAMED: {history} -> {desired}")
                else:
                    for f_name in os.listdir(history_path):
                        srcf = os.path.join(history_path, f_name)
                        dstf = os.path.join(desired_path, f_name)
                        if os.path.exists(dstf) and os.path.isfile(srcf):
                            safe_copy(srcf, dstf)
                            try:
                                os.remove(srcf)
                            except OSError:
                                pass
                        elif not os.path.exists(dstf):
                            shutil.move(srcf, dstf)
                    try:
                        os.rmdir(history_path)
                    except OSError:
                        pass
                df.loc[df["history folder name"] == history, "history folder name"] = desired
                processed_folders[key] = True
            if not os.path.exists(desired_path):
                os.makedirs(desired_path, exist_ok=True)
            df.at[idx, "history folder name"] = desired
        except Exception as e:
            log_msg(f"ERROR row {idx}: {e}")
            log_error(f"P1 ERROR row {idx}: {e}")

    for idx, row in tqdm(df.iterrows(), total=len(df), desc="[P1] Copying ISOs", ncols=80):
        iso_no = row["Iso no"].strip()
        folder = row["folder name"].strip()
        dest_folder = os.path.join(dest_root, folder)
        if not folder or not iso_no:
            continue
        src_iso = find_iso_on_server(iso_no, server_path)
        if not src_iso:
            os.makedirs(dest_folder, exist_ok=True)
            df.at[idx, "ISO Status"] = "MISSING"
            log_error(f"P1: MISSING ISO {iso_no}")
            continue
        dest_iso = os.path.join(dest_folder, os.path.basename(src_iso))
        try:
            safe_copy(src_iso, dest_iso)
            df.at[idx, "ISO Status"] = "OK"
        except Exception as e:
            log_msg(f"ERROR copying {iso_no}: {e}")
            log_error(f"P1: ERROR copying {iso_no}: {e}")
            df.at[idx, "ISO Status"] = "MISSING"

    df.to_excel(excel_file, index=False)
    highlight_missing_iso(excel_file)
    elapsed = time.time() - process_start
    log_msg(f"=== P1: ISO MANAGER END ({format_time(elapsed)}) ===\n")

# --- PROCESS 2: GENERATE EXCEL ---
def generate_excel(dest_root):
    process_start = time.time()
    log_msg("=== P2: GENERATE EXCEL START ===")
    pdf_pattern = re.compile(r'\(([^)]+)\)\.pdf$', re.IGNORECASE)
    processed = 0
    
    # CRITICAL FIX: Only process direct children of dest_root
    for root, dirs, _ in os.walk(dest_root):
        if root == dest_root:
            for dir_name in tqdm(dirs, desc="[P2] Generating Excel", ncols=80):
                subdir = os.path.join(root, dir_name)
                excel_path = os.path.join(subdir, OUTPUT_EXCEL_NAME)
                iso_set = set()
                
                pdf_files = [f for f in os.listdir(subdir) if f.lower().endswith(FINAL_PDF_SUFFIX)]
                
                for file in pdf_files:
                    match = pdf_pattern.search(file)
                    if match:
                        parenthesis_content = match.group(1)
                        segments = parenthesis_content.split('-')
                        if len(segments) >= 2:
                            iso = f"{segments[0].strip()}-{segments[1].strip()}"
                            iso_set.add(iso)
                
                if iso_set:
                    existing_iso = set()
                    if os.path.exists(excel_path):
                        try:
                            # Read existing ISOs only if file exists
                            df_existing = pd.read_excel(excel_path, dtype=str).fillna("")
                            existing_iso = set(df_existing.get("ISO LIST", pd.Series()).dropna().astype(str))
                        except Exception as e:
                            log_error(f"P2: Error reading existing Excel in {subdir}: {e}")
                            
                    new_isos = [iso for iso in sorted(iso_set) if iso not in existing_iso]
                    
                    if new_isos or not os.path.exists(excel_path):
                        df_new = pd.DataFrame({"ISO LIST": new_isos})
                        if existing_iso and os.path.exists(excel_path):
                            # Retain any non-ISO LIST columns from existing file if it exists
                            df_final = pd.concat([df_existing.filter(items=["ISO LIST"]), df_new], ignore_index=True).drop_duplicates()
                        else:
                            df_final = df_new
                        
                        df_final.to_excel(excel_path, index=False)
                        processed += 1
                        log_msg(f"Updated Excel: {subdir}")
                # Note on Empty Directory Handling (Error 7): Folders without PDFs are correctly skipped.
    
    elapsed = time.time() - process_start
    log_msg(f"=== P2: GENERATE EXCEL END ({format_time(elapsed)}) - {processed} folders ===\n")

# --- PROCESS 3: EXTRACT PAGES (EXTERNAL INDEX) ---
def extract_pages(dest_root, pdf_path, page_index_excel):
    process_start = time.time()
    log_msg("=== P3: EXTRACT PAGES START (Reading External Index) ===")
    
    if not os.path.exists(pdf_path):
        log_msg(f"ERROR: Master PDF not found: {pdf_path}")
        log_error(f"P3: Master PDF not found {pdf_path}")
        return
    
    if not os.path.exists(page_index_excel):
        log_msg(f"ERROR: Page Index Excel not found: {page_index_excel}. Skipping P3.")
        log_error(f"P3: Page Index Excel not found {page_index_excel}")
        return

    # 1. BUILD GLOBAL ISO -> PAGE MAP from the Master Index Excel
    iso_page_map = {}
    try:
        master_df = pd.read_excel(page_index_excel, dtype=str).fillna("")
        if 'ISO LIST' not in master_df.columns or 'PDF PAGE' not in master_df.columns:
            log_msg("ERROR: Master Index Excel requires columns 'ISO LIST' and 'PDF PAGE'. Skipping P3.")
            log_error("P3: Master Index Excel format error.")
            return

        for _, row in master_df.iterrows():
            iso = str(row['ISO LIST']).strip()
            pages = str(row['PDF PAGE']).strip()
            if iso and pages:
                iso_page_map[iso.upper()] = pages
        
        log_msg(f"Loaded {len(iso_page_map)} ISO entries from Master Page Index.")

    except Exception as e:
        log_msg(f"FATAL ERROR reading Master Index Excel: {e}")
        log_error(f"P3: Error reading Master Index: {e}")
        return

    # 2. PERFORM EXTRACTION
    extracted = 0
    try:
        pdf_reader = PyPDF2.PdfReader(pdf_path)
        total_pages = len(pdf_reader.pages)
    except Exception as e:
        log_msg(f"ERROR: Cannot read Master PDF {e}")
        log_error(f"P3: Cannot read Master PDF {e}")
        return
    
    # CRITICAL FIX: Only process direct children of dest_root
    for root, dirs, _ in os.walk(dest_root):
        if root == dest_root:
            for dir_name in tqdm(dirs, desc="[P3] Extracting Pages", ncols=80):
                subdir = os.path.join(root, dir_name)
                excel_path = os.path.join(subdir, OUTPUT_EXCEL_NAME)
                
                if not os.path.exists(excel_path):
                    continue
                
                try:
                    # CRITICAL FIX: Improved error handling for reading the system Excel
                    system_df = pd.read_excel(excel_path, dtype=str)
                    if "ISO LIST" not in system_df.columns:
                        log_error(f"P3: 'ISO LIST' column missing in {excel_path}. Skipping folder.")
                        continue
                    
                    for iso_list_entry in system_df["ISO LIST"].dropna():
                        iso = str(iso_list_entry).strip().upper()
                        if not iso:
                            continue
                        
                        pages_str = iso_page_map.get(iso, "")
                        
                        if pages_str:
                            for p in pages_str.split(','):
                                p = p.strip()
                                if p.isdigit():
                                    page_num = int(p)
                                    
                                    if 1 <= page_num <= total_pages:
                                        output_path = os.path.join(subdir, f"{page_num}{FINAL_PDF_SUFFIX}")
                                        if not os.path.exists(output_path):
                                            try:
                                                pdf_writer = PyPDF2.PdfWriter()
                                                pdf_writer.add_page(pdf_reader.pages[page_num - 1]) 
                                                with open(output_path, 'wb') as f:
                                                    pdf_writer.write(f)
                                                extracted += 1
                                            except Exception as e:
                                                log_error(f"P3: Error extracting page {page_num} for ISO {iso}: {e}")
                                    else:
                                        log_error(f"P3: Invalid page number {page_num} for ISO {iso} (Total pages: {total_pages}).")
                except Exception as e:
                    log_error(f"P3: Error processing {subdir} or system Excel: {e}")
    
    elapsed = time.time() - process_start
    log_msg(f"=== P3: EXTRACT PAGES END ({format_time(elapsed)}) - {extracted} pages ===\n")

# --- PROCESS 4: FRICOPY (CACHING) ---
def extract_name_in_parentheses(filename):
    match = re.search(r"\(([^)]+)\)", filename)
    return match.group(1).strip() if match else None

def get_dir_hash(path):
    total_size = 0
    file_count = 0
    for dirpath, _, filenames in os.walk(path):
        for f in filenames:
            full_path = os.path.join(dirpath, f)
            if f.lower().endswith(FINAL_PDF_SUFFIX): 
                try:
                    total_size += os.path.getsize(full_path)
                    file_count += 1
                except:
                    pass
    return hashlib.sha256(f"{file_count}-{total_size}".encode()).hexdigest()

def create_linewise_index_cached(root_path):
    current_hash = get_dir_hash(root_path)

    if os.path.exists(P4_CACHE_FILE):
        try:
            with open(P4_CACHE_FILE, 'r') as f:
                cache_data = json.load(f)
            
            if cache_data.get('hash') == current_hash:
                log_msg("Loaded P4 index from cache (LINEWISE unchanged).")
                indexed = {k: [(p, n) for p, n in v] for k, v in cache_data.get('index', {})}
                return indexed
            else:
                log_msg("P4 index cache invalid (LINEWISE changed or structure mismatch). Rebuilding...")
        except Exception as e:
            log_msg(f"ERROR loading cache: {e}. Rebuilding...")

    linewise_index = defaultdict(list)
    for subdir, _, files in tqdm(os.walk(root_path), desc="[P4] Building Index", ncols=80, leave=False):
        for linewise_file in files:
            if linewise_file.lower().endswith(FINAL_PDF_SUFFIX):
                search_key = os.path.splitext(linewise_file)[0].lower()
                full_path = os.path.join(subdir, linewise_file)
                linewise_index[search_key].append((full_path, linewise_file))

    try:
        cache_index = {k: [[p, n] for p, n in v] for k, v in linewise_index.items()}
        with open(P4_CACHE_FILE, 'w') as f:
            json.dump({'hash': current_hash, 'index': cache_index}, f)
        log_msg("P4 index rebuilt and saved to cache.")
    except Exception as e:
        log_msg(f"ERROR saving P4 cache: {e}")
        
    return linewise_index

def fricopy(backup_root, linewise_root):
    process_start = time.time()
    log_msg("=== P4: FRICOPY START (Using Cached Index) ===")
    
    linewise_index = create_linewise_index_cached(linewise_root)
    if not linewise_index:
        log_msg("ERROR: Failed to index LINEWISE")
        log_error("P4: Failed to index LINEWISE")
        return
    
    all_backup_files = []
    # CRITICAL FIX: Only process direct children of dest_root
    for root, dirs, _ in os.walk(backup_root):
        if root == backup_root:
            for dir_name in dirs:
                subdir = os.path.join(root, dir_name)
                for file in os.listdir(subdir):
                    if file.lower().endswith(FINAL_PDF_SUFFIX) and not file.lower().endswith("_fri.pdf"):
                        all_backup_files.append((subdir, file))

    copied = 0
    for subdir, file in tqdm(all_backup_files, desc="[P4] Creating FRI copies", ncols=80):
        base_name = extract_name_in_parentheses(file)
        if not base_name:
            continue
        base_name_lower = base_name.lower()
        
        for linewise_key, linewise_entries in linewise_index.items():
            if base_name_lower in linewise_key:
                for full_path, lw_file in linewise_entries:
                    target_filename = os.path.splitext(lw_file)[0] + "_FRI.pdf"
                    target_path = os.path.join(subdir, target_filename)
                    if not os.path.exists(target_path) or os.path.getsize(full_path) != os.path.getsize(target_path):
                        try:
                            shutil.copy2(full_path, target_path)
                            copied += 1
                        except Exception as e:
                            log_error(f"P4: Error copying FRI: {e}")
    
    elapsed = time.time() - process_start
    log_msg(f"=== P4: FRICOPY END ({format_time(elapsed)}) - {copied} copies ===\n")

# --- PROCESS 5: CLEANUP REDUNDANCY ---
def cleanup_redundancy(dest_root, excel_file):
    process_start = time.time()
    log_msg("=== P5: CLEANUP REDUNDANCY START ===")
    
    df = pd.read_excel(excel_file, dtype=str).fillna("")
    folder_iso_map = {}
    for _, row in df.iterrows():
        folder = row["folder name"].strip()
        iso_no = row["Iso no"].strip()
        if folder and iso_no:
            if folder not in folder_iso_map:
                folder_iso_map[folder] = []
            folder_iso_map[folder].append(iso_no)
    
    deleted = 0
    # CRITICAL FIX: Only process direct children of dest_root (os.listdir is appropriate here)
    for folder_name in tqdm(os.listdir(dest_root), desc="[P5] Cleaning redundancy", ncols=80):
        folder_path = os.path.join(dest_root, folder_name)
        if not os.path.isdir(folder_path):
            continue
        
        expected_isos = folder_iso_map.get(folder_name, [])
        if not expected_isos:
            continue
        
        for file_name in os.listdir(folder_path):
            file_path = os.path.join(folder_path, file_name)
            if not os.path.isfile(file_path):
                continue
            
            # Keep the final combined PDF
            if file_name.lower() == f"{folder_name.lower()}{FINAL_PDF_SUFFIX}":
                continue

            belongs = False
            # Check if file name contains any expected ISO
            for iso_no in expected_isos:
                if iso_no.lower() in file_name.lower():
                    belongs = True
                    break
            
            # Check for extracted page PDFs (e.g., '1.pdf', '10.pdf')
            if re.match(r'^\d+\.pdf$', file_name.lower()):
                 belongs = True
            
            # Check for FRI copies
            if file_name.lower().endswith("_fri.pdf"):
                belongs = True

            if not belongs:
                try:
                    os.remove(file_path)
                    deleted += 1
                except Exception as e:
                    log_error(f"P5: Error deleting {file_name}: {e}")
    
    elapsed = time.time() - process_start
    log_msg(f"=== P5: CLEANUP REDUNDANCY END ({format_time(elapsed)}) - {deleted} deleted ===\n")

# --- PROCESS 6: COMBINE PDF (CUSTOM SORTING & CACHING) ---
def page_hash(page):
    return hashlib.sha256(page.get_text("xml").encode('utf-8')).hexdigest()

def get_folder_source_hash(folder_path, pdf_files):
    current_hash = hashlib.md5()
    for f in sorted(pdf_files):
        fpath = os.path.join(folder_path, f)
        try:
            with open(fpath, 'rb') as file:
                current_hash.update(file.read())
        except Exception as e:
            log_error(f"P6: Error hashing {f}: {e}")
            return None 
    return current_hash.hexdigest()

def get_sort_key(filename):
    """
    Defines the custom sort order: Numeric -> ISO -> FRI
    """
    name = filename.lower()
    base, _ = os.path.splitext(name)
    
    # 1. Numeric PDFs: <0, number> (e.g., 1.pdf, 10.pdf)
    if re.fullmatch(r'\d+', base):
        try:
            return (0, int(base)) 
        except ValueError:
            return (99, base)

    # 3. FRI ISO PDFs: <2, filename> (e.g., (123-456)_FRI.pdf)
    elif name.endswith("_fri.pdf"):
        return (2, name.replace("_fri.pdf", FINAL_PDF_SUFFIX))

    # 2. Main ISO PDFs: <1, filename> (e.g., (123-456).pdf)
    else:
        return (1, name)

def combine_pdfs(dest_root):
    process_start = time.time()
    log_msg("=== P6: COMBINE PDF START (Using Merge Cache and Custom Sort) ===")
    
    combined = 0
    # CRITICAL FIX: Only process direct children of dest_root
    for root, dirs, _ in os.walk(dest_root):
        if root == dest_root:
            for dir_name in tqdm(dirs, desc="[P6] Combining PDFs", ncols=80):
                folder_path = os.path.join(root, dir_name)
                output_pdf_path = os.path.join(folder_path, f"{dir_name}{FINAL_PDF_SUFFIX}")
                cache_file = os.path.join(folder_path, P6_CACHE_NAME)
                
                # 1. Filter out the final combined PDF
                pdf_files = [
                    f for f in os.listdir(folder_path) 
                    if f.lower().endswith(FINAL_PDF_SUFFIX) and f.lower() != f"{dir_name.lower()}{FINAL_PDF_SUFFIX}" and not f.startswith('.')
                ]
                
                if not pdf_files:
                    continue

                # Check for changes (Reconstruction Logic)
                current_source_hash = get_folder_source_hash(folder_path, pdf_files)
                
                if os.path.exists(cache_file):
                    try:
                        with open(cache_file, 'r') as f:
                            cache_data = json.load(f)
                        
                        if cache_data.get('source_hash') == current_source_hash:
                            continue # Skip reconstruction
                    except:
                        pass
                
                # 2. Apply Custom Sort
                pdf_files.sort(key=get_sort_key)
                
                combined_pdf = fitz.open()
                seen_hashes = set()
                
                try:
                    for filename in pdf_files:
                        try:
                            full_file_path = os.path.join(folder_path, filename)
                            with fitz.open(full_file_path) as pdf:
                                for page in pdf:
                                    # Deduplication logic
                                    ph = page_hash(page)
                                    if ph not in seen_hashes:
                                        combined_pdf.insert_pdf(pdf, from_page=page.number, to_page=page.number)
                                        seen_hashes.add(ph)
                        except Exception as e:
                            log_error(f"P6: Error combining {filename}: {e}")
                    
                    # 3. Save the combined file
                    if combined_pdf.page_count > 0:
                        combined_pdf.save(output_pdf_path, garbage=4, deflate=True)
                        combined += 1
                        
                        # Update cache
                        try:
                            with open(cache_file, 'w') as f:
                                json.dump({'source_hash': current_source_hash, 'timestamp': datetime.now().isoformat()}, f)
                        except Exception as e:
                            log_error(f"P6: Error writing cache for {dir_name}: {e}")
                            
                finally:
                    if 'combined_pdf' in locals() and combined_pdf:
                        combined_pdf.close()
    
    elapsed = time.time() - process_start
    log_msg(f"=== P6: COMBINE PDF END ({format_time(elapsed)}) - {combined} combined ===\n")

# --- PROCESS 7: FINAL CLEANUP + VERIFY ---
def final_cleanup_and_verify(dest_root, excel_file):
    process_start = time.time()
    log_msg("=== P7: FINAL CLEANUP + ERROR CHECK START ===")
    
    df = pd.read_excel(excel_file, dtype=str).fillna("")
    issues_found = 0
    
    # CRITICAL FIX: Only process direct children of dest_root (os.listdir is appropriate here)
    for folder_name in os.listdir(dest_root):
        folder_path = os.path.join(dest_root, folder_name)
        if not os.path.isdir(folder_path):
            continue
        
        folder_in_excel = any(df["folder name"] == folder_name)
        
        # Check for orphaned empty folders not in the main Excel
        if not folder_in_excel and not os.listdir(folder_path):
            try:
                os.rmdir(folder_path)
                log_msg(f"Deleted orphaned empty folder: {folder_name}")
            except:
                pass
        
        # Check for folders in Excel that are now empty (issue)
        if folder_in_excel and not os.listdir(folder_path):
            log_error(f"P7: Empty folder {folder_name} listed in main Excel.")
            issues_found += 1
    
    missing_count = len(df[df["ISO Status"] == "MISSING"])
    ok_count = len(df[df["ISO Status"] == "OK"])
    
    report = f"""
================================================================================
                        FINAL VERIFICATION REPORT
================================================================================
Total ISOs processed: {len(df)}
ISOs OK: {ok_count}
ISOs MISSING: {missing_count}
Issues Found (e.g., empty folders in main Excel): {issues_found}

Excel file: {excel_file}
Destination: {dest_root}

All errors logged in: {ERROR_REPORT}
================================================================================
"""
    
    print(report)
    log_msg(report)
    
    elapsed = time.time() - process_start
    log_msg(f"=== P7: FINAL CLEANUP + ERROR CHECK END ({format_time(elapsed)}) ===\n")
    
    return missing_count, issues_found

# ============================================================================
# MAIN ORCHESTRATOR
# ============================================================================
def main():
    master_start = time.time()
    
    # Initialize logs
    with open(LOG_FILE, "w", encoding="utf-8") as f:
        f.write("=== ORCHESTRATOR LOG ===\n\n")
    with open(ERROR_REPORT, "w", encoding="utf-8") as f:
        f.write("=== ERROR REPORT ===\n\n")
    
    log_msg("Starting Complete Orchestrator...\n")
    
    # The primary Excel mapping loop/system to ISO.
    excel_file = os.path.join(os.getcwd(), "loop_system_iso.xlsx")
    excel_dir = os.path.dirname(excel_file)
    if not create_or_update_excel(excel_file):
        return
    
    print("\n" + "="*80)
    print("SELECT REQUIRED FOLDERS AND FILES (7 PROCESS WORKFLOW):")
    print("="*80)
    
    # --- USER INPUT SELECTIONS ---
    print("FOLDER 1: Select the master repository for original ISOs.")
    server_path = select_folder("FOLDER 1: Server ISO Folder (Source for P1)")
    if not server_path:
        log_msg("Workflow cancelled: Server ISO Folder not selected.")
        return
    
    print("FOLDER 2: Select the root destination for ALL generated folders/files.")
    dest_root = select_folder("FOLDER 2: Destination Root (Workspace)")
    if not dest_root:
        log_msg("Workflow cancelled: Destination Root not selected.")
        return
    
    print("FOLDER 3: Select the LINEWISE archive folder (Source for P4/FRI copies).")
    linewise_path = select_folder("FOLDER 3: LINEWISE Folder (Source for P4)")
    if not linewise_path:
        log_msg("Workflow cancelled: LINEWISE Folder not selected.")
        return
    
    print("FILE 4: Select the MASTER PDF file (Source document for page extraction in P3).")
    pdf_path = select_file("FILE 4: Master PDF (Source for P3)", [("PDF files", "*.pdf")])
    if not pdf_path:
        log_msg("Workflow cancelled: Master PDF not selected.")
        return
        
    print("**FILE 5: Select the PRE-GENERATED PDF INDEX (.txt file) for user reference.**")
    pdf_index_txt = select_file("FILE 5: PDF Index (.txt) Reference", [("Text files", "*.txt"), ("All files", "*.*")])
    if not pdf_index_txt:
        log_msg("Workflow cancelled: PDF Index (.txt) Reference not selected.")
        return
    
    print("**FILE 6: Select the MASTER PAGE INDEX EXCEL file (ISO to Page Number map for P3).**")
    page_index_excel = select_file("FILE 6: Master Page Index Excel (ISO to Page Number map for P3)", [("Excel files", "*.xlsx")])
    if not page_index_excel:
        log_msg("Workflow cancelled: Master Page Index Excel not selected.")
        return
        
    # --- Step 1: Copy the .txt index file to the working directory for user convenience ---
    try:
        shutil.copy2(pdf_index_txt, os.path.join(excel_dir, PDF_INDEX_REFERENCE_NAME))
        log_msg(f"Copied PDF Index Reference to: {os.path.join(excel_dir, PDF_INDEX_REFERENCE_NAME)}")
        messagebox.showinfo("Index Reference Ready", 
                            f"The PDF Index Reference has been copied to the script folder ({PDF_INDEX_REFERENCE_NAME}).\n\n"
                            f"Please ensure the Master Page Index Excel ({os.path.basename(page_index_excel)}) has been fully updated "
                            f"with page numbers from this reference *before* proceeding.")
    except Exception as e:
        log_msg(f"ERROR copying PDF Index Reference: {e}")
        messagebox.showerror("Error", f"Failed to copy PDF Index Reference. Please check permissions. Workflow cancelled.")
        return

    summary = f"""
P1: ISO Manager (Source: {os.path.basename(server_path)})
P2: Generate Excel (Workspace: {os.path.basename(dest_root)})
P3: Extract Pages (Source PDF: {os.path.basename(pdf_path)})
    - Uses Page Index: {os.path.basename(page_index_excel)}
    - Reference File: {PDF_INDEX_REFERENCE_NAME} (copied to script folder)
P4: Fricopy (Source: {os.path.basename(linewise_path)})
P5: Cleanup Redundancy
P6: Combine PDF (Custom Sorted & Cached)
P7: Final Cleanup + Verify
"""
    print(summary)
    
    if not messagebox.askyesno("Confirm", "Run complete 7-process workflow?"):
        return
    
    # --- EXECUTE WORKFLOW ---
    try:
        iso_manager(excel_file, server_path, dest_root)
        generate_excel(dest_root)
        
        # P3: Extraction based on external ISO-Page mapping
        extract_pages(dest_root, pdf_path, page_index_excel)
        
        # P4: Copy FRI documents
        fricopy(dest_root, linewise_path)
        
        # P5: Delete unnecessary files
        cleanup_redundancy(dest_root, excel_file)
        
        # P6: Combine PDFs with custom sorting and caching
        combine_pdfs(dest_root) 
        
        # P7: Final checks
        missing, issues = final_cleanup_and_verify(dest_root, excel_file)
        
        total_elapsed = time.time() - master_start
        
        result = f"""
================================================================================
COMPLETE WORKFLOW FINISHED
================================================================================
Total Time: {format_time(total_elapsed)}
Missing ISOs: {missing}
Issues Found: {issues}

Logs:
- Main: {LOG_FILE}
- Errors: {ERROR_REPORT}
================================================================================
"""
        print(result)
        messagebox.showinfo("Success", result)
    except Exception as e:
        log_msg(f"FATAL ERROR in main execution: {e}")
        log_error(f"FATAL: {e}")
        messagebox.showerror("Error", f"Failed: {e}\n\nCheck error report.")

if __name__ == "__main__":
    main()
