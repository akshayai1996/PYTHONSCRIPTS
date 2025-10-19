"""
===============================================================================
 Project Name:    ISO Folder Manager with GUI (v2.8)
 File Name:       Iso_Folder_Manager_GUI.py
 Description:     
    Manages ISO files and folders based on Excel master sheet.
    Key Features:
    1. Excel columns: Iso no | loop no | system no | folder name | history folder name | ISO Status
    2. Missing ISOs create placeholder file ISO_NOT_FOUND.txt
    3. Rows with missing ISOs highlighted in red in Excel
    4. GUI folder selection for server ISO folder and destination root folder
    5. Handles added, updated, or deleted rows
    6. Prevents duplicate folder creation
    7. Logs all actions in folder_action_log.txt
    8. ADDED PROGRESS TRACKING using tqdm
 Author:          Akshay Solanki
 Created On:      19-Oct-2025
 Dependencies:    pandas, openpyxl, tkinter, os, shutil, datetime, tqdm
===============================================================================
"""

import os
import shutil
import pandas as pd
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tqdm import tqdm

def log(msg: str, log_file):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{timestamp}] {msg}"
    print(line)
    try:
        with open(log_file, "a", encoding="utf-8") as f:
            f.write(line + "\n")
    except Exception as e:
        print(f"ERROR: Could not write to log file: {e}")

def safe_copy(src, dest):
    if not os.path.exists(src):
        raise FileNotFoundError(src)
    dest_dir = os.path.dirname(dest)
    os.makedirs(dest_dir, exist_ok=True)
    base, ext = os.path.splitext(dest)
    if os.path.exists(dest) and os.path.getsize(src) == os.path.getsize(dest):
        return dest
    candidate = dest
    i = 1
    while os.path.exists(candidate):
        candidate = f"{base}_dup{i}{ext}"
        i += 1
    shutil.copy2(src, candidate)
    return candidate

def make_folder_name(loop_no: str, system_no: str) -> str:
    loop_no = str(loop_no).strip()
    system_no = str(system_no).strip()
    return f"{loop_no}_{system_no}"

def find_iso_on_server(iso_no: str, server_path: str) -> str:
    if not iso_no or not isinstance(iso_no, str):
        return ""
    candidate = os.path.join(server_path, iso_no)
    if os.path.isfile(candidate):
        return candidate
    if os.path.isdir(server_path):
        try:
            for item in os.listdir(server_path):
                if item.lower() == iso_no.lower():
                    full_path = os.path.join(server_path, item)
                    if os.path.isfile(full_path):
                        return full_path
        except PermissionError:
            pass
    return ""

def create_or_update_excel(excel_file, log_file):
    headers = ["Iso no", "loop no", "system no", "folder name", "history folder name", "ISO Status"]
    
    if not os.path.exists(excel_file):
        df = pd.DataFrame(columns=headers)
        df.to_excel(excel_file, index=False)
        log(f"Created Excel: {excel_file}", log_file)
        messagebox.showinfo("Excel Created", f"Excel created:\n{excel_file}\n\nFill first three columns (Iso no, loop no, system no).")
        return
        
    df = pd.read_excel(excel_file, dtype=str).fillna("")
    for col in headers:
        if col not in df.columns:
            df[col] = ""
    df = df[headers]
    for idx, row in df.iterrows():
        new_folder_name = make_folder_name(row["loop no"], row["system no"])
        df.at[idx, "folder name"] = new_folder_name
        if not row["history folder name"]:
            df.at[idx, "history folder name"] = new_folder_name
    df.to_excel(excel_file, index=False)
    log(f"Excel normalized: {excel_file}", log_file)
    messagebox.showinfo("Excel Ready", f"Excel ready:\n{excel_file}")

def highlight_missing_iso(excel_file, log_file):
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        clear_fill = PatternFill(fill_type=None)
        
        header = [cell.value for cell in ws[1]]
        if "ISO Status" not in header:
            return
        
        col_idx = header.index("ISO Status") + 1
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            status_val = str(cell.value).strip().upper() if cell.value else ""
            fill = red_fill if status_val == "MISSING" else clear_fill
            for c in range(1, ws.max_column + 1):
                ws.cell(row=row, column=c).fill = fill
        
        wb.save(excel_file)
        log("Excel highlighting applied.", log_file)
    except Exception as e:
        log(f"ERROR applying highlighting: {e}", log_file)

def sync_folders_and_copy(excel_file, server_path, dest_root, log_file):
    if not os.path.exists(excel_file):
        messagebox.showerror("Error", "Excel not found.")
        return
        
    df = pd.read_excel(excel_file, dtype=str).fillna("")
    for idx, row in df.iterrows():
        df.at[idx, "folder name"] = make_folder_name(row["loop no"], row["system no"])

    processed_folders = {}
    
    log("--- Starting Folder Structure Sync (Phase 1/2) ---", log_file)
    for idx, row in df.iterrows():
        desired = row["folder name"].strip()
        history = row["history folder name"].strip()
        if not desired or not history:
            continue
        
        key = (history, desired)
        if key in processed_folders:
            df.at[idx, "history folder name"] = desired
            continue
        
        desired_path = os.path.join(dest_root, desired)
        history_path = os.path.join(dest_root, history)
        
        try:
            if history != desired and os.path.exists(history_path):
                if not os.path.exists(desired_path):
                    os.rename(history_path, desired_path)
                    log(f"RENAMED: {history} -> {desired}", log_file)
                else:
                    for f_name in os.listdir(history_path):
                        srcf = os.path.join(history_path, f_name)
                        dstf = os.path.join(desired_path, f_name)
                        if os.path.exists(dstf) and os.path.isfile(srcf):
                            safe_copy(srcf, dstf)
                            try:
                                os.remove(srcf)
                            except OSError:
                                pass
                        elif not os.path.exists(dstf):
                            shutil.move(srcf, dstf)
                            log(f"MOVED {f_name}: {history} -> {desired}", log_file)
                    
                    try:
                        os.rmdir(history_path)
                        log(f"DELETED empty folder: {history}", log_file)
                    except OSError:
                        log(f"Folder not empty: {history}", log_file)
                
                df.loc[df["history folder name"] == history, "history folder name"] = desired
                processed_folders[key] = True

            if not os.path.exists(desired_path):
                os.makedirs(desired_path, exist_ok=True)
                log(f"CREATED folder: {desired}", log_file)
            
            df.at[idx, "history folder name"] = desired
        except Exception as e:
            log(f"ERROR row {idx}: {e}", log_file)

    log("--- Starting ISO Copy Phase (Phase 2/2) ---", log_file)
    for idx, row in tqdm(df.iterrows(), total=len(df), desc="Copying ISOs", ncols=80):
        iso_no = row["Iso no"].strip()
        folder = row["folder name"].strip()
        dest_folder = os.path.join(dest_root, folder)
        
        if not folder or not iso_no:
            continue
        
        src_iso = find_iso_on_server(iso_no, server_path)
        
        if not src_iso:
            os.makedirs(dest_folder, exist_ok=True)
            with open(os.path.join(dest_folder, "ISO_NOT_FOUND.txt"), "w") as f:
                f.write(f"ISO {iso_no} not found.\n")
            df.at[idx, "ISO Status"] = "MISSING"
            log(f"MISSING ISO: {iso_no} (row {idx})", log_file)
            continue
        
        dest_iso = os.path.join(dest_folder, os.path.basename(src_iso))
        try:
            safe_copy(src_iso, dest_iso)
            df.at[idx, "ISO Status"] = "OK"
            log(f"COPIED: {iso_no} -> {folder}", log_file)
        except Exception as e:
            log(f"ERROR copying {iso_no}: {e}", log_file)
            df.at[idx, "ISO Status"] = "MISSING"

    log("--- Starting Destination Cleanup ---", log_file)
    try:
        existing = set(os.listdir(dest_root))
        excel_folders = set(df["folder name"].tolist())
        for folder in existing:
            path = os.path.join(dest_root, folder)
            if folder not in excel_folders and os.path.isdir(path) and not os.listdir(path):
                try:
                    os.rmdir(path)
                    log(f"DELETED empty: {folder}", log_file)
                except Exception as e:
                    log(f"ERROR deleting {folder}: {e}", log_file)
    except Exception as e:
        log(f"ERROR cleanup: {e}", log_file)

    df.to_excel(excel_file, index=False)
    highlight_missing_iso(excel_file, log_file)
    messagebox.showinfo("Done", "Sync complete. Check log for details.")

def select_folder(title="Select folder"):
    root = tk.Tk()
    root.withdraw()
    folder = filedialog.askdirectory(title=title)
    root.destroy()
    return folder

def main():
    log_file = os.path.join(os.getcwd(), "folder_action_log.txt")
    excel_file = os.path.join(os.getcwd(), "loop_system_iso.xlsx")
    
    create_or_update_excel(excel_file, log_file)
    server_path = select_folder("Select Server ISO Folder")
    if not server_path:
        log("Server folder cancelled.", log_file)
        return
    dest_root = select_folder("Select Destination Root")
    if not dest_root:
        log("Destination cancelled.", log_file)
        return
    if not messagebox.askyesno("Confirm", "Proceed with sync? This may take time."):
        log("Cancelled by user.", log_file)
        return
    
    sync_folders_and_copy(excel_file, server_path, dest_root, log_file)

if __name__ == "__main__":
    main()